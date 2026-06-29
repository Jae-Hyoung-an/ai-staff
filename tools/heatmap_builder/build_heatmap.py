# -*- coding: utf-8 -*-
"""
상점 위치/주문수 CSV -> 인터랙티브 오더 히트맵(단일 HTML) 빌더.

- 행정동 경계: vuski/admdongkor (EPSG:4326 GeoJSON)
- 자치구(구) 경계: southkorea/seoul-maps
- 두 경계는 데이터 bbox에 걸치는 영역만 잘라(clip) HTML에 인라인 임베드.
- 히트맵은 Leaflet + leaflet.heat 기반, 외부 서버 없이 더블클릭으로 실행 가능.

사용법:
    python build_heatmap.py
    python build_heatmap.py --csv "경로/데이터.csv" --out "경로/index.html" --title "성북구 오더 히트맵"

자세한 내용은 같은 폴더의 README.md 참고.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    import requests
except ImportError:
    print("[오류] requests 가 필요합니다. `pip install -r requirements.txt` 를 먼저 실행하세요.")
    sys.exit(1)


# --- 경계 데이터 소스 ---------------------------------------------------------
ADM_URL = (
    "https://raw.githubusercontent.com/vuski/admdongkor/master/"
    "ver20260401/HangJeongDong_ver20260401.geojson"
)
GU_URL = (
    "https://raw.githubusercontent.com/southkorea/seoul-maps/master/"
    "juso/2015/json/seoul_municipalities_geo_simple.json"
)

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent
CACHE_DIR = SCRIPT_DIR / ".cache"

DEFAULT_CSVS = [
    Path(r"c:\Users\jaehyoung.an\Downloads\성북구_데이터.csv"),
    Path(r"c:\Users\jaehyoung.an\Downloads\260618_성북성동기존오더.csv"),
    Path(r"c:\Users\jaehyoung.an\Downloads\성동구데이터.xlsx"),
]
DEFAULT_OUT = REPO_ROOT / "projects" / "260617_성북구_오더_히트맵" / "index.html"

BBOX_MARGIN = 0.012  # 약 1.2km 여유

# 컬럼명 별칭 (파일마다 헤더가 다를 수 있음)
NAME_KEYS = ("상호명", "상점명", "리스트_상호명", "사이트_상호명", "가맹점명", "상점", "이름")
ORDER_KEYS = ("최근주문수", "오더수", "주문수")
DELIVER_KEYS = ("가게배달유무", "가게배달여부", "가게배달", "배달유무")
PRICE_KEYS = ("가게배달가격", "가게배달 가격", "배달가격", "가게배달비")

# 상호명 부분일치 폴백: 별칭에 없어도 '상호명/상점명' 등이 들어간 컬럼을 자동 인식
NAME_CONTAINS = ("상호명", "상점명", "상호", "상점명", "가맹점")

# 가게배달유무 카테고리 코드
#   Y = 가게배달 가능, N = 불가, E = 기존오더
def _deliver_code(raw: str) -> str:
    v = (raw or "").strip()
    if v.upper() == "Y":
        return "Y"
    if v.upper() == "N":
        return "N"
    if v == "기존오더":
        return "E"
    return v or "N"


# --- 유틸 --------------------------------------------------------------------
def log(msg: str) -> None:
    print(msg, flush=True)


def download(url: str, cache_name: str) -> dict:
    """GeoJSON 다운로드(캐시 사용)."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = CACHE_DIR / cache_name
    if cache_path.exists() and cache_path.stat().st_size > 0:
        log(f"  - 캐시 사용: {cache_name}")
        return json.loads(cache_path.read_text(encoding="utf-8"))
    log(f"  - 다운로드 중: {url}")
    resp = requests.get(url, headers={"User-Agent": "heatmap-builder"}, timeout=120)
    resp.raise_for_status()
    resp.encoding = "utf-8"
    cache_path.write_text(resp.text, encoding="utf-8")
    return resp.json()


def _pick(row: dict, keys: tuple) -> str:
    for k in keys:
        v = row.get(k)
        if v not in (None, ""):
            return v
    return ""


def _pick_name(row: dict) -> str:
    """상호명: 별칭 우선, 없으면 컬럼명 부분일치로 폴백."""
    v = _pick(row, NAME_KEYS)
    if v:
        return v
    for key, val in row.items():
        if val in (None, ""):
            continue
        if any(tok in str(key) for tok in NAME_CONTAINS):
            return val
    return ""


def _load_rows(path: Path) -> list[dict]:
    """CSV/XLSX를 dict 행 목록으로 로드."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("xlsx 읽기에는 openpyxl 이 필요합니다. `pip install -r requirements.txt`")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(it)]
        rows = []
        for vals in it:
            if vals is None or all(v is None for v in vals):
                continue
            rows.append({header[i]: vals[i] for i in range(min(len(header), len(vals)))})
        log(f"  - xlsx 시트 '{wb.sheetnames[0]}' 에서 {len(rows)}행 읽음")
        return rows
    # CSV: 인코딩 자동 판별
    last_err = None
    for enc in ("utf-8-sig", "cp949", "utf-8"):
        try:
            with open(path, encoding=enc, newline="") as f:
                rows = list(csv.DictReader(f))
            log(f"  - 인코딩 {enc} 로 {len(rows)}행 읽음")
            return rows
        except (UnicodeDecodeError, LookupError) as e:
            last_err = e
    raise RuntimeError(f"CSV 인코딩 판별 실패: {last_err}")


def read_csv(csv_path: Path) -> list[dict]:
    """CSV/XLSX 읽고 검증된 상점 목록 반환. 컬럼명은 별칭 허용."""
    rows = _load_rows(csv_path)

    stores = []
    skipped = 0
    for r in rows:
        try:
            lat = float(r["위도"])
            lon = float(r["경도"])
            order = int(float(_pick(r, ORDER_KEYS) or 0))
        except (ValueError, TypeError, KeyError):
            skipped += 1
            continue
        # 한반도 대략 범위 밖이면 제외
        if not (33.0 <= lat <= 39.5 and 124.0 <= lon <= 132.0):
            skipped += 1
            continue
        store = {
            "n": str(_pick_name(r)).strip(),
            "la": round(lat, 6),
            "lo": round(lon, 6),
            "d": _deliver_code(str(_pick(r, DELIVER_KEYS))),
            "o": order,
        }
        price = str(_pick(r, PRICE_KEYS)).strip()
        if price:
            store["p"] = price  # 가게배달 가격 (있을 때만 저장 -> 용량 절약)
        stores.append(store)
    if skipped:
        log(f"  - 좌표/형식 이상으로 {skipped}행 제외")
    if not stores:
        raise RuntimeError("유효한 상점 데이터가 없습니다.")
    return stores


def data_bbox(stores: list[dict]) -> tuple[float, float, float, float]:
    lats = [s["la"] for s in stores]
    lons = [s["lo"] for s in stores]
    return (
        min(lons) - BBOX_MARGIN,
        min(lats) - BBOX_MARGIN,
        max(lons) + BBOX_MARGIN,
        max(lats) + BBOX_MARGIN,
    )


def _iter_points(coords):
    """중첩 좌표 배열에서 [lon, lat] 쌍을 순회."""
    if (
        len(coords) >= 2
        and isinstance(coords[0], (int, float))
        and isinstance(coords[1], (int, float))
    ):
        yield coords
        return
    for c in coords:
        yield from _iter_points(c)


def feature_bbox(geom: dict):
    xs, ys = [], []
    for lon, lat in _iter_points(geom["coordinates"]):
        xs.append(lon)
        ys.append(lat)
    if not xs:
        return None
    return (min(xs), min(ys), max(xs), max(ys))


def bbox_intersects(a, b) -> bool:
    return not (a[2] < b[0] or a[0] > b[2] or a[3] < b[1] or a[1] > b[3])


def round_coords(coords, ndigits=5):
    """좌표 정밀도를 낮춰 파일 용량 축소."""
    if (
        len(coords) >= 2
        and isinstance(coords[0], (int, float))
        and isinstance(coords[1], (int, float))
    ):
        return [round(coords[0], ndigits), round(coords[1], ndigits)]
    return [round_coords(c, ndigits) for c in coords]


def clip_geojson(gj: dict, bbox, name_key: str, keep_props: list[str]) -> dict:
    """bbox에 걸치는 feature만 남기고 속성/좌표를 경량화."""
    out_features = []
    for feat in gj.get("features", []):
        geom = feat.get("geometry")
        if not geom or "coordinates" not in geom:
            continue
        fb = feature_bbox(geom)
        if fb is None or not bbox_intersects(fb, bbox):
            continue
        props = feat.get("properties", {})
        new_props = {k: props.get(k) for k in keep_props if k in props}
        new_props["_name"] = props.get(name_key, "")
        out_features.append(
            {
                "type": "Feature",
                "properties": new_props,
                "geometry": {
                    "type": geom["type"],
                    "coordinates": round_coords(geom["coordinates"]),
                },
            }
        )
    return {"type": "FeatureCollection", "features": out_features}


# --- HTML 빌드 ---------------------------------------------------------------
def build_html(stores, adm_gj, gu_gj, bbox, title: str) -> str:
    center_lat = (bbox[1] + bbox[3]) / 2
    center_lon = (bbox[0] + bbox[2]) / 2
    max_order = max((s["o"] for s in stores), default=1) or 1

    config = {
        "title": title,
        "center": [round(center_lat, 6), round(center_lon, 6)],
        "bounds": [[bbox[1], bbox[0]], [bbox[3], bbox[2]]],
        "maxOrder": max_order,
    }

    safe_title = title.replace("<", "").replace(">", "")
    template = _HTML_TEMPLATE
    template = template.replace("__TITLE_TAG__", safe_title)
    template = template.replace("/*__CONFIG__*/", json.dumps(config, ensure_ascii=False))
    template = template.replace("/*__STORES__*/", json.dumps(stores, ensure_ascii=False, separators=(",", ":")))
    template = template.replace("/*__ADM__*/", json.dumps(adm_gj, ensure_ascii=False, separators=(",", ":")))
    template = template.replace("/*__GU__*/", json.dumps(gu_gj, ensure_ascii=False, separators=(",", ":")))
    return template


def main() -> None:
    ap = argparse.ArgumentParser(description="상점 CSV -> 인터랙티브 오더 히트맵 HTML")
    ap.add_argument("--csv", nargs="+", default=[str(p) for p in DEFAULT_CSVS],
                    help="입력 CSV 경로(여러 개 가능, 공백으로 구분)")
    ap.add_argument("--out", default=str(DEFAULT_OUT), help="출력 HTML 경로")
    ap.add_argument("--title", default="성북·성동 오더 히트맵", help="지도 제목")
    args = ap.parse_args()

    out_path = Path(args.out)
    csv_paths = [Path(c) for c in args.csv]
    missing = [str(p) for p in csv_paths if not p.exists()]
    if missing:
        log(f"[오류] CSV 파일을 찾을 수 없습니다: {', '.join(missing)}")
        sys.exit(1)

    log("[1/4] CSV 읽는 중...")
    stores = []
    for p in csv_paths:
        log(f"  · {p.name}")
        stores.extend(read_csv(p))
    bbox = data_bbox(stores)
    n_y = sum(1 for s in stores if s["d"] == "Y")
    n_n = sum(1 for s in stores if s["d"] == "N")
    n_e = sum(1 for s in stores if s["d"] == "E")
    log(f"  - 상점 {len(stores)}개 (배달 Y={n_y}, N={n_n}, 기존오더={n_e})")
    log(f"  - 영역 bbox: {tuple(round(v, 4) for v in bbox)}")

    log("[2/4] 행정동 경계 처리 중...")
    adm_raw = download(ADM_URL, "hangjeongdong.geojson")
    adm_gj = clip_geojson(adm_raw, bbox, name_key="adm_nm", keep_props=["adm_nm", "sggnm"])
    log(f"  - 행정동 {len(adm_gj['features'])}개 추출")

    log("[3/4] 자치구 경계 처리 중...")
    gu_raw = download(GU_URL, "seoul_municipalities.geojson")
    gu_gj = clip_geojson(gu_raw, bbox, name_key="SIG_KOR_NM", keep_props=["SIG_KOR_NM"])
    log(f"  - 자치구 {len(gu_gj['features'])}개 추출")

    log("[4/4] HTML 생성 중...")
    html = build_html(stores, adm_gj, gu_gj, bbox, args.title)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")
    size_mb = out_path.stat().st_size / 1e6
    log(f"  - 완료: {out_path} ({size_mb:.2f} MB)")
    log("\n브라우저에서 위 HTML 파일을 더블클릭하면 바로 열립니다.")


_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>__TITLE_TAG__</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
<style>
  :root { --panel-bg: rgba(255,255,255,0.96); --accent:#e8511d; }
  html, body { margin:0; padding:0; height:100%; font-family: "Malgun Gothic","맑은 고딕",system-ui,sans-serif; }
  #map { position:absolute; inset:0; }
  .panel {
    position:absolute; top:12px; right:12px; z-index:1000;
    background:var(--panel-bg); border-radius:12px; padding:14px 16px;
    box-shadow:0 4px 18px rgba(0,0,0,0.18); width:236px; font-size:13px; color:#222;
    max-height:calc(100vh - 24px); overflow-y:auto;
  }
  .phead { display:flex; align-items:center; justify-content:space-between; gap:8px; }
  .panel h1 { font-size:15px; margin:0 0 4px; }
  .ptoggle { display:none; border:none; background:transparent; font-size:20px; line-height:1;
    cursor:pointer; color:#555; padding:2px 4px; border-radius:6px; }
  .ptoggle:hover { background:#0001; }
  .panel.collapsed { width:auto; }
  .panel.collapsed #panelBody { display:none; }
  .panel.collapsed h1 { margin:0; }
  .panel .sub { font-size:11px; color:#777; margin:0 0 12px; }
  .panel .group { border-top:1px solid #eee; padding:10px 0 4px; }
  .panel .group:first-of-type { border-top:none; }
  .panel .label { font-weight:600; font-size:12px; color:#555; margin-bottom:6px; display:block; }
  .panel label.opt { display:flex; align-items:center; gap:7px; padding:3px 0; cursor:pointer; }
  .panel input { cursor:pointer; }
  .count { font-weight:700; color:var(--accent); }
  .legend { position:absolute; bottom:18px; left:18px; z-index:1000;
    background:var(--panel-bg); border-radius:10px; padding:10px 12px; font-size:12px;
    box-shadow:0 4px 18px rgba(0,0,0,0.15); }
  .legend .bar { height:10px; width:160px; border-radius:5px; margin:6px 0 4px;
    background:linear-gradient(to right,#2b6cff,#36e0c8,#7cf04a,#f5e13a,#f59f1b,#e8201a); }
  .legend .ends { display:flex; justify-content:space-between; color:#666; }
  .swatch { display:inline-block; width:14px; height:0; border-top:3px solid; margin-right:6px; vertical-align:middle; }
  .leaflet-popup-content { font-size:13px; line-height:1.5; }
  .toggle-collapse { position:absolute; top:12px; right:12px; z-index:1001; display:none; }
  /* 행정동 이름 라벨 */
  .adm-label {
    background:transparent; border:none; box-shadow:none; padding:0;
    color:#7a2a10; font-size:11px; font-weight:700; white-space:nowrap;
    text-shadow:0 0 3px #fff,0 0 3px #fff,0 0 3px #fff,0 0 3px #fff;
  }
  .adm-label::before { display:none; } /* 말풍선 꼬리 제거 */
  .leaflet-container.labels-off .adm-label { display:none; }
  /* 모바일 대응 */
  @media (max-width: 640px) {
    .panel { top:10px; right:10px; width:min(76vw, 250px); padding:11px 13px; font-size:13px; }
    .ptoggle { display:block; }
    .panel label.opt { padding:6px 0; } /* 터치 영역 확대 */
    .legend { bottom:12px; left:12px; padding:8px 10px; font-size:11px; }
    .legend .bar { width:130px; }
    .adm-label { font-size:10px; }
    .leaflet-control-attribution { font-size:9px; }
  }
</style>
</head>
<body>
<div id="map"></div>

<div class="panel" id="panel">
  <div class="phead">
    <h1 id="ptitle">오더 히트맵</h1>
    <button id="panelToggle" class="ptoggle" aria-label="설정 패널 열기/닫기" title="설정 열기/닫기">&#9776;</button>
  </div>
  <div id="panelBody">
  <p class="sub">표시 상점 <span class="count" id="cnt">0</span>개 · 주문수 기준 히트맵</p>

  <div class="group">
    <span class="label">히트맵 강도 기준</span>
    <label class="opt"><input type="radio" name="weight" value="order" checked /> 최근주문수 가중</label>
    <label class="opt"><input type="radio" name="weight" value="density" /> 상점 밀도(균등)</label>
  </div>

  <div class="group">
    <span class="label">표시 대상</span>
    <label class="opt"><input type="radio" name="filter" value="all" checked /> 전체 오더</label>
    <label class="opt"><input type="radio" name="filter" value="baemin" /> 배민 자료 전체 (Y/N)</label>
    <label class="opt"><input type="radio" name="filter" value="Y" /> 가게배달만 (Y)</label>
    <label class="opt"><input type="radio" name="filter" value="E" /> 기존오더만</label>
  </div>

  <div class="group">
    <span class="label">경계 표시</span>
    <label class="opt"><input type="checkbox" id="chkAdm" checked /> <span class="swatch" style="border-color:#e8511d"></span>행정동 경계</label>
    <label class="opt"><input type="checkbox" id="chkGu" /> <span class="swatch" style="border-color:#1d4ed8"></span>자치구(구) 경계</label>
    <label class="opt"><input type="checkbox" id="chkStore" checked /> 상점 위치 점 표시</label>
  </div>
  </div><!-- /panelBody -->
</div>

<div class="legend">
  <div>주문 밀집도 (낮음 → 높음)</div>
  <div class="bar"></div>
  <div class="ends"><span>낮음</span><span>높음</span></div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.heat@0.2.0/dist/leaflet-heat.js"></script>
<script>
const CONFIG = /*__CONFIG__*/;
const STORES = /*__STORES__*/;
const ADM = /*__ADM__*/;
const GU = /*__GU__*/;

document.getElementById('ptitle').textContent = CONFIG.title;
document.title = CONFIG.title;

const map = L.map('map', { preferCanvas:true }).setView(CONFIG.center, 13);
L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png', {
  attribution:'&copy; OpenStreetMap, &copy; CARTO',
  subdomains:'abcd', maxZoom:19
}).addTo(map);
try { map.fitBounds(CONFIG.bounds, { padding:[20,20] }); } catch(e) {}

const gradient = {0.0:'#2b6cff',0.25:'#36e0c8',0.45:'#7cf04a',0.65:'#f5e13a',0.82:'#f59f1b',1.0:'#e8201a'};
let heat = L.heatLayer([], { radius:26, blur:20, maxZoom:17, minOpacity:0.25, max:1.0, gradient:gradient }).addTo(map);

function matchFilter(s, filt) {
  if (filt === 'all') return true;
  if (filt === 'baemin') return s.d === 'Y' || s.d === 'N';
  return s.d === filt;
}

function buildPoints() {
  const mode = document.querySelector('input[name=weight]:checked').value;
  const filt = document.querySelector('input[name=filter]:checked').value;
  const rows = STORES.filter(s => matchFilter(s, filt));
  // 현재 표시 대상 기준으로 정규화 -> 각 보기에서 색 대비 최대화
  const mx = rows.reduce((m, s) => Math.max(m, s.o), 0) || 1;
  const pts = rows.map(s => {
    const w = (mode === 'order') ? Math.max(0.05, s.o / mx) : 0.5;
    return [s.la, s.lo, w];
  });
  heat.setLatLngs(pts);
  document.getElementById('cnt').textContent = rows.length.toLocaleString('ko-KR');
}

// 경계 레이어
function shortName(nm){ const p=(nm||'').split(' '); return p[p.length-1] || ''; }
const admLayer = L.geoJSON(ADM, {
  style:{ color:'#e8511d', weight:1.6, fill:true, fillColor:'#e8511d', fillOpacity:0.03 },
  onEachFeature:(f,l)=> l.bindTooltip(shortName(f.properties._name), {
    permanent:true, direction:'center', className:'adm-label', opacity:1
  })
});
// 줌이 낮으면 라벨이 겹치므로 숨김
const LABEL_MIN_ZOOM = 13;
function updateLabelVisibility(){
  const el = map.getContainer();
  if (map.getZoom() < LABEL_MIN_ZOOM) el.classList.add('labels-off');
  else el.classList.remove('labels-off');
}
map.on('zoomend', updateLabelVisibility);
const guLayer = L.geoJSON(GU, {
  style:{ color:'#1d4ed8', weight:2.4, fill:false, dashArray:'5,4' },
  onEachFeature:(f,l)=> l.bindTooltip(f.properties._name || '', {sticky:true})
});

// 상점 점 레이어 (요청 시)
let storeLayer = null;
const DELIVER_STYLE = {
  Y: { stroke:'#0a8f5b', fill:'#13c47d', label:'Y (가게배달 가능)', color:'#0a8f5b' },
  N: { stroke:'#888',    fill:'#bbb',    label:'N (가게배달 불가)', color:'#b04b00' },
  E: { stroke:'#6a35c9', fill:'#9b6be6', label:'기존오더',          color:'#6a35c9' }
};
function dstyle(d) { return DELIVER_STYLE[d] || DELIVER_STYLE.N; }

function storeInfoHtml(s) {
  const st = dstyle(s.d);
  return `<div style="min-width:160px">
       <div style="font-size:14px;font-weight:700;margin-bottom:6px">${s.n || '(이름없음)'}</div>
       <div>구분: <b style="color:${st.color}">${st.label}</b></div>
       <div>주문수: <b>${s.o.toLocaleString('ko-KR')}</b> 건</div>
       <div>가게배달 가격: <b>${s.p || ''}</b></div>
     </div>`;
}

function buildStoreLayer() {
  const filt = document.querySelector('input[name=filter]:checked').value;
  const rows = STORES.filter(s => matchFilter(s, filt));
  return L.layerGroup(rows.map(s => {
    const st = dstyle(s.d);
    const html = storeInfoHtml(s);
    return L.circleMarker([s.la, s.lo], {
      radius:4, color:st.stroke, weight:1, fillColor:st.fill, fillOpacity:0.85
    }).bindTooltip(html, { direction:'top', offset:[0,-4], opacity:0.96 })
      .bindPopup(html);
  }));
}
function refreshStoreLayer() {
  const on = document.getElementById('chkStore').checked;
  if (storeLayer) { map.removeLayer(storeLayer); storeLayer = null; }
  if (on) { storeLayer = buildStoreLayer(); storeLayer.addTo(map); }
}

// 이벤트
document.querySelectorAll('input[name=weight]').forEach(el => el.addEventListener('change', buildPoints));
document.querySelectorAll('input[name=filter]').forEach(el => el.addEventListener('change', () => { buildPoints(); refreshStoreLayer(); }));
document.getElementById('chkAdm').addEventListener('change', e => e.target.checked ? admLayer.addTo(map) : map.removeLayer(admLayer));
document.getElementById('chkGu').addEventListener('change', e => e.target.checked ? guLayer.addTo(map) : map.removeLayer(guLayer));
document.getElementById('chkStore').addEventListener('change', refreshStoreLayer);

// 모바일: 패널 접기/펼치기
const panel = document.getElementById('panel');
document.getElementById('panelToggle').addEventListener('click', () => panel.classList.toggle('collapsed'));
if (window.matchMedia('(max-width: 640px)').matches) panel.classList.add('collapsed');

// 행정동 경계는 기본 표시
if (document.getElementById('chkAdm').checked) admLayer.addTo(map);

buildPoints();
refreshStoreLayer();
updateLabelVisibility();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
